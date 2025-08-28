from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_wtf.csrf import CSRFProtect
import mysql.connector
import bcrypt
import os
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from functools import wraps
import logging

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production-2024-protrack-rpt-system')
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = 3600
csrf = CSRFProtect(app)

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'protrack_rpt'
}

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_db_connection():
    """Create and return a database connection"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except mysql.connector.Error as err:
        logger.error(f"Database connection error: {err}")
        return None

def init_database():
    """Initialize database tables if they don't exist"""
    connection = get_db_connection()
    if not connection:
        return False
    
    cursor = connection.cursor()
    
    try:
        # Create consumables table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS consumables (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                description TEXT,
                category VARCHAR(100),
                quantity INT DEFAULT 0,
                damaged INT DEFAULT 0,
                returnable TINYINT(1) DEFAULT 1,
                image_url VARCHAR(500),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Ensure new columns exist on legacy deployments
        try:
            cursor.execute("ALTER TABLE consumables ADD COLUMN IF NOT EXISTS damaged INT DEFAULT 0")
        except mysql.connector.Error:
            pass
        try:
            cursor.execute("ALTER TABLE consumables ADD COLUMN IF NOT EXISTS returnable TINYINT(1) DEFAULT 1")
        except mysql.connector.Error:
            pass
        
        # Create orders table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_name VARCHAR(255) NOT NULL,
                department VARCHAR(255) NOT NULL,
                purpose TEXT NOT NULL,
                date_needed DATE NOT NULL,
                status ENUM('Pending', 'Approved', 'Rejected') DEFAULT 'Pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create order_items table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INT AUTO_INCREMENT PRIMARY KEY,
                order_id INT,
                consumable_id INT,
                quantity INT NOT NULL,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
            )
        """)

        # Create consumable borrow and return tables
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS consumable_borrows (
                id INT AUTO_INCREMENT PRIMARY KEY,
                consumable_id INT NOT NULL,
                borrower_name VARCHAR(255) NOT NULL,
                borrower_type ENUM('Student','Staff') NOT NULL,
                contact_info VARCHAR(255),
                department VARCHAR(255),
                quantity INT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (consumable_id) REFERENCES consumables(id) ON DELETE CASCADE
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS consumable_returns (
                id INT AUTO_INCREMENT PRIMARY KEY,
                borrow_id INT NOT NULL,
                returned_quantity INT DEFAULT 0,
                damaged_quantity INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (borrow_id) REFERENCES consumable_borrows(id) ON DELETE CASCADE
            )
        """)
        
        # Create admin_users table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS admin_users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create audit_logs table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(100),
                action VARCHAR(255),
                details TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Create laboratory table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS laboratory (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                status ENUM('Active', 'Inactive', 'Maintenance') DEFAULT 'Active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create lab_assets table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS lab_assets (
                id INT AUTO_INCREMENT PRIMARY KEY,
                lab_id INT NOT NULL,
                name VARCHAR(255) NOT NULL,
                category VARCHAR(100) NOT NULL,
                status ENUM('Available', 'In Use', 'Maintenance', 'Retired') DEFAULT 'Available',
                purchase_date DATE,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (lab_id) REFERENCES laboratory(id) ON DELETE CASCADE
            )
        """)
        # Add asset_code column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE lab_assets ADD COLUMN IF NOT EXISTS asset_code VARCHAR(100)")
        except mysql.connector.Error:
            pass
        # Add Damaged to status enum if needed
        try:
            cursor.execute("ALTER TABLE lab_assets MODIFY COLUMN status ENUM('Available','In Use','Maintenance','Retired','Damaged') DEFAULT 'Available'")
        except mysql.connector.Error:
            pass
        # Rename purchase_date -> stock_date if present
        try:
            cursor.execute("ALTER TABLE lab_assets CHANGE COLUMN purchase_date stock_date DATE")
        except mysql.connector.Error:
            # Fallback: ensure stock_date exists and copy data from purchase_date if both exist
            try:
                cursor.execute("ALTER TABLE lab_assets ADD COLUMN IF NOT EXISTS stock_date DATE")
                cursor.execute("UPDATE lab_assets SET stock_date = purchase_date WHERE stock_date IS NULL")
            except mysql.connector.Error:
                pass
        
        # Create asset_categories table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS asset_categories (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(100) UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Insert default admin user if not exists
        cursor.execute("SELECT COUNT(*) FROM admin_users")
        if cursor.fetchone()[0] == 0:
            hashed_password = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt())
            cursor.execute("""
                INSERT INTO admin_users (username, password) VALUES (%s, %s)
            """, ('admin', hashed_password.decode('utf-8')))
            
            # Insert sample consumables
            sample_items = [
                ('Office Paper A4', 'High quality A4 paper for printing', 'Office Supplies', 500, '/static/images/paper.jpg'),
                ('Blue Pens', 'Blue ballpoint pens, pack of 10', 'Writing Supplies', 100, '/static/images/pens.jpg'),
                ('Stapler', 'Heavy duty stapler with staples', 'Office Equipment', 25, '/static/images/stapler.jpg'),
                ('Notebooks', 'Spiral bound notebooks, A5 size', 'Writing Supplies', 75, '/static/images/notebooks.jpg'),
                ('USB Cables', 'USB Type-C cables, 1m length', 'Electronics', 50, '/static/images/usb.jpg')
            ]
            
            for item in sample_items:
                cursor.execute("""
                    INSERT INTO consumables (name, description, category, quantity, image_url) 
                    VALUES (%s, %s, %s, %s, %s)
                """, item)

        # Seed default asset categories if table is empty
        cursor.execute("SELECT COUNT(*) FROM asset_categories")
        if cursor.fetchone()[0] == 0:
            default_categories = [
                ('Computer Equipment',),
                ('Lab Equipment',),
                ('Furniture',),
                ('Electronics',),
                ('Tools',)
            ]
            cursor.executemany("INSERT INTO asset_categories (name) VALUES (%s)", default_categories)
        
        connection.commit()
        return True
        
    except mysql.connector.Error as err:
        logger.error(f"Database initialization error: {err}")
        return False
    finally:
        cursor.close()
        connection.close()

def admin_required(f):
    """Decorator to require admin login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please login as admin to access this page', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def log_admin_action(action, details):
    """Log admin actions for audit trail"""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor()
        try:
            cursor.execute("""
                INSERT INTO audit_logs (admin_username, action, details) 
                VALUES (%s, %s, %s)
            """, (session.get('admin_username'), action, details))
            connection.commit()
        except mysql.connector.Error as err:
            logger.error(f"Audit log error: {err}")
        finally:
            cursor.close()
            connection.close()

@app.route('/')
def index():
    """Public home page with consumables listing"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('index.html', consumables=[])
    
    cursor = connection.cursor(dictionary=True)
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    page = request.args.get('page', 1, type=int)
    per_page = 12
    
    # Build query with filters
    query = "SELECT * FROM consumables WHERE 1=1"
    params = []
    
    if search:
        query += " AND (name LIKE %s OR description LIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    if category:
        query += " AND category = %s"
        params.append(category)
    
    # Get total count for pagination
    count_query = query.replace("SELECT *", "SELECT COUNT(*)")
    cursor.execute(count_query, params)
    total_items = cursor.fetchone()['COUNT(*)']
    
    # Add pagination
    query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
    params.extend([per_page, (page - 1) * per_page])
    
    cursor.execute(query, params)
    consumables = cursor.fetchall()
    
    # Get unique categories for filter
    cursor.execute("SELECT DISTINCT category FROM consumables ORDER BY category")
    categories = [row['category'] for row in cursor.fetchall()]
    
    cursor.close()
    connection.close()
    
    total_pages = (total_items + per_page - 1) // per_page
    
    return render_template('index.html', 
                         consumables=consumables, 
                         categories=categories,
                         search=search,
                         category=category,
                         page=page,
                         total_pages=total_pages)

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    """Add item to cart (session-based)"""
    if 'cart' not in session:
        session['cart'] = {}
    
    consumable_id = request.form.get('consumable_id')
    quantity = int(request.form.get('quantity', 1))
    
    if consumable_id in session['cart']:
        session['cart'][consumable_id] += quantity
    else:
        session['cart'][consumable_id] = quantity
    
    session.modified = True
    flash('Item added to cart successfully!', 'success')
    return redirect(url_for('index'))

@app.route('/cart')
def cart():
    """View cart contents"""
    if 'cart' not in session or not session['cart']:
        flash('Your cart is empty', 'info')
        return redirect(url_for('index'))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('index'))
    
    cursor = connection.cursor(dictionary=True)
    cart_items = []
    total = 0
    
    for consumable_id, quantity in session['cart'].items():
        cursor.execute("SELECT * FROM consumables WHERE id = %s", (consumable_id,))
        item = cursor.fetchone()
        if item:
            item['cart_quantity'] = quantity
            cart_items.append(item)
            total += quantity
    
    cursor.close()
    connection.close()
    
    return render_template('cart.html', cart_items=cart_items, total=total)

@app.route('/update_cart', methods=['POST'])
def update_cart():
    """Update cart quantities"""
    consumable_id = request.form.get('consumable_id')
    quantity = int(request.form.get('quantity', 0))
    
    if quantity <= 0:
        session['cart'].pop(consumable_id, None)
    else:
        session['cart'][consumable_id] = quantity
    
    session.modified = True
    flash('Cart updated successfully!', 'success')
    return redirect(url_for('cart'))

@app.route('/place_order', methods=['GET', 'POST'])
def place_order():
    """Place order from cart"""
    if request.method == 'GET':
        if 'cart' not in session or not session['cart']:
            flash('Your cart is empty', 'info')
            return redirect(url_for('index'))
        return render_template('place_order.html')
    
    # Process order
    user_name = request.form.get('user_name')
    department = request.form.get('department')
    purpose = request.form.get('purpose')
    date_needed = request.form.get('date_needed')
    
    if not all([user_name, department, purpose, date_needed]):
        flash('Please fill in all required fields', 'error')
        return render_template('place_order.html')
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('place_order.html')
    
    cursor = connection.cursor()
    
    try:
        # Create order
        cursor.execute("""
            INSERT INTO orders (user_name, department, purpose, date_needed) 
            VALUES (%s, %s, %s, %s)
        """, (user_name, department, purpose, date_needed))
        
        order_id = cursor.lastrowid
        
        # Add order items
        for consumable_id, quantity in session['cart'].items():
            cursor.execute("""
                INSERT INTO order_items (order_id, consumable_id, quantity) 
                VALUES (%s, %s, %s)
            """, (order_id, consumable_id, quantity))
        
        connection.commit()
        
        # Clear cart
        session.pop('cart', None)
        
        flash('Order placed successfully! Your order ID is: ' + str(order_id), 'success')
        return redirect(url_for('index'))
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash('Error placing order. Please try again.', 'error')
        logger.error(f"Order placement error: {err}")
    finally:
        cursor.close()
        connection.close()
    
    return render_template('place_order.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        connection = get_db_connection()
        if not connection:
            flash('Database connection error', 'error')
            return render_template('admin/login.html')
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM admin_users WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
            session['admin_logged_in'] = True
            session['admin_username'] = username
            log_admin_action('Login', f'Admin {username} logged in')
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    if 'admin_logged_in' in session:
        log_admin_action('Logout', f'Admin {session.get("admin_username")} logged out')
        session.pop('admin_logged_in', None)
        session.pop('admin_username', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard with statistics"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('admin/dashboard.html')
    
    cursor = connection.cursor(dictionary=True)
    
    # Get statistics
    cursor.execute("SELECT COUNT(*) as total_items FROM consumables")
    total_items = cursor.fetchone()['total_items']
    
    cursor.execute("SELECT COUNT(*) as total_orders FROM orders")
    total_orders = cursor.fetchone()['total_orders']
    
    cursor.execute("SELECT COUNT(*) as pending_orders FROM orders WHERE status = 'Pending'")
    pending_orders = cursor.fetchone()['pending_orders']
    
    cursor.execute("SELECT COUNT(*) as low_stock FROM consumables WHERE quantity < 10")
    low_stock = cursor.fetchone()['low_stock']
    
    # Get recent orders
    cursor.execute("""
        SELECT o.*, COUNT(oi.id) as item_count 
        FROM orders o 
        LEFT JOIN order_items oi ON o.id = oi.order_id 
        GROUP BY o.id 
        ORDER BY o.created_at DESC 
        LIMIT 5
    """)
    recent_orders = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return render_template('admin/dashboard.html',
                         total_items=total_items,
                         total_orders=total_orders,
                         pending_orders=pending_orders,
                         low_stock=low_stock,
                         recent_orders=recent_orders)

@app.route('/admin/inventory')
@admin_required
def admin_inventory():
    """Manage laboratories list"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('admin/inventory.html', labs=[], search='', sort_by='name', sort_order='asc', page=1, total_pages=0)

    cursor = connection.cursor(dictionary=True)
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Build query with search
    query = "SELECT * FROM laboratory WHERE 1=1"
    params = []
    
    if search:
        query += " AND (name LIKE %s OR status LIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    # Get total count for pagination
    count_query = query.replace("SELECT *", "SELECT COUNT(*)")
    cursor.execute(count_query, params)
    total_labs = cursor.fetchone()['COUNT(*)']
    
    # Add sorting and pagination
    query += f" ORDER BY {sort_by} {sort_order.upper()}"
    query += " LIMIT %s OFFSET %s"
    params.extend([per_page, (page - 1) * per_page])
    
    try:
        cursor.execute(query, params)
        labs = cursor.fetchall()
    finally:
        cursor.close()
        connection.close()

    total_pages = (total_labs + per_page - 1) // per_page
    
    return render_template('admin/inventory.html', 
                         labs=labs, 
                         search=search,
                         sort_by=sort_by,
                         sort_order=sort_order,
                         page=page,
                         total_pages=total_pages)


@app.route('/admin/consumables')
@admin_required
def admin_consumables():
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('admin/consumables.html', items=[], search='')

    search = request.args.get('search', '')
    cursor = connection.cursor(dictionary=True)
    try:
        params = []
        query = """
            SELECT c.*,
                   COALESCE((SELECT COALESCE(SUM(b.quantity),0) - COALESCE(SUM(r.returned_quantity + r.damaged_quantity),0)
                     FROM consumable_borrows b LEFT JOIN consumable_returns r ON r.borrow_id = b.id
                     WHERE b.consumable_id = c.id),0) AS borrowed
            FROM consumables c
            WHERE 1=1
        """
        if search:
            query += " AND (c.name LIKE %s OR c.category LIKE %s)"
            params.extend([f'%{search}%', f'%{search}%'])
        query += " ORDER BY c.created_at DESC"
        cursor.execute(query, params)
        items = cursor.fetchall()
    finally:
        cursor.close()
        connection.close()

    return render_template('admin/consumables.html', items=items, search=search)


@app.route('/admin/consumables/add', methods=['POST'])
@admin_required
def admin_consumables_add():
    name = (request.form.get('name') or '').strip()
    quantity = int(request.form.get('quantity') or 0)
    stock_date = request.form.get('stock_date') or None
    category = (request.form.get('category') or '').strip()
    returnable = 1 if (request.form.get('returnable') == 'on' or request.form.get('returnable') == '1') else 0

    if not name or quantity <= 0 or not category:
        flash('Please provide Name, positive Quantity, and Category', 'error')
        return redirect(url_for('admin_consumables'))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_consumables'))

    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO consumables (name, category, quantity, returnable)
            VALUES (%s, %s, %s, %s)
            """,
            (name, category, quantity, returnable)
        )
        connection.commit()
        flash('Consumable added successfully', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Add consumable error: {err}")
        flash('Error adding consumable', 'error')
    finally:
        cursor.close()
        connection.close()
    return redirect(url_for('admin_consumables'))


def _consumable_has_records(cursor, consumable_id: int) -> bool:
    cursor.execute("SELECT COUNT(*) AS cnt FROM consumable_borrows WHERE consumable_id = %s", (consumable_id,))
    return (cursor.fetchone() or {}).get('cnt', 0) > 0


@app.route('/admin/consumables/<int:cid>/update', methods=['POST'])
@admin_required
def admin_consumables_update(cid):
    name = (request.form.get('name') or '').strip()
    quantity = int(request.form.get('quantity') or 0)
    category = (request.form.get('category') or '').strip()
    returnable = 1 if (request.form.get('returnable') == 'on' or request.form.get('returnable') == '1') else 0

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_consumables'))

    cursor = connection.cursor(dictionary=True)
    try:
        if _consumable_has_records(cursor, cid):
            flash('Cannot update: consumable has borrow/provide records', 'error')
            return redirect(url_for('admin_consumables'))
        cursor = connection.cursor()
        cursor.execute(
            """
            UPDATE consumables SET name=%s, quantity=%s, category=%s, returnable=%s WHERE id=%s
            """,
            (name, quantity, category, returnable, cid)
        )
        connection.commit()
        flash('Consumable updated', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Update consumable error: {err}")
        flash('Error updating consumable', 'error')
    finally:
        cursor.close()
        connection.close()
    return redirect(url_for('admin_consumables'))


@app.route('/admin/consumables/<int:cid>/delete', methods=['POST'])
@admin_required
def admin_consumables_delete(cid):
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_consumables'))

    cursor = connection.cursor(dictionary=True)
    try:
        if _consumable_has_records(cursor, cid):
            flash('Cannot delete: consumable has borrow/provide records', 'error')
            return redirect(url_for('admin_consumables'))
        cursor = connection.cursor()
        cursor.execute("DELETE FROM consumables WHERE id=%s", (cid,))
        connection.commit()
        flash('Consumable deleted', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Delete consumable error: {err}")
        flash('Error deleting consumable', 'error')
    finally:
        cursor.close()
        connection.close()
    return redirect(url_for('admin_consumables'))


@app.route('/admin/consumables/<int:cid>/borrow', methods=['POST'])
@admin_required
def admin_consumables_borrow(cid):
    borrower_name = (request.form.get('borrower_name') or '').strip()
    borrower_type = (request.form.get('borrower_type') or '').strip()
    contact_info = (request.form.get('contact_info') or '').strip()
    department = (request.form.get('department') or '').strip()
    quantity = int(request.form.get('quantity') or 0)

    if not borrower_name or borrower_type not in ('Student','Staff') or quantity <= 0:
        flash('Provide valid borrower info and positive quantity', 'error')
        return redirect(url_for('admin_consumables'))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_consumables'))

    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, quantity, returnable FROM consumables WHERE id=%s", (cid,))
        cons = cursor.fetchone()
        if not cons:
            flash('Consumable not found', 'error')
            return redirect(url_for('admin_consumables'))
        if cons['quantity'] < quantity:
            flash('Insufficient stock', 'error')
            return redirect(url_for('admin_consumables'))

        # Deduct stock
        cursor2 = connection.cursor()
        cursor2.execute("UPDATE consumables SET quantity = quantity - %s WHERE id=%s", (quantity, cid))
        # Create borrow
        cursor2.execute(
            """
            INSERT INTO consumable_borrows (consumable_id, borrower_name, borrower_type, contact_info, department, quantity)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (cid, borrower_name, borrower_type, contact_info, department, quantity)
        )
        connection.commit()
        flash('Borrow recorded and stock updated', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Borrow error: {err}")
        flash('Error recording borrow', 'error')
    finally:
        cursor.close()
        connection.close()
    return redirect(url_for('admin_consumables'))


@app.route('/admin/consumables/returns/<int:borrow_id>', methods=['POST'])
@admin_required
def admin_consumables_return(borrow_id):
    returned_quantity = int(request.form.get('returned_quantity') or 0)
    damaged_quantity = int(request.form.get('damaged_quantity') or 0)

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_consumables'))

    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT b.*, c.returnable FROM consumable_borrows b JOIN consumables c ON c.id=b.consumable_id WHERE b.id=%s", (borrow_id,))
        b = cursor.fetchone()
        if not b:
            flash('Borrow record not found', 'error')
            return redirect(url_for('admin_consumables'))
        if not b['returnable']:
            flash('Item is non-returnable', 'error')
            return redirect(url_for('admin_consumables'))

        total = returned_quantity + damaged_quantity
        if total != b['quantity']:
            flash('Returned + Damaged must equal Borrowed quantity', 'error')
            return redirect(url_for('admin_consumables'))

        # Record return
        cursor2 = connection.cursor()
        cursor2.execute(
            """
            INSERT INTO consumable_returns (borrow_id, returned_quantity, damaged_quantity)
            VALUES (%s, %s, %s)
            """,
            (borrow_id, returned_quantity, damaged_quantity)
        )
        # Update stock: add returned to quantity; add damaged to damaged
        cursor2.execute(
            "UPDATE consumables SET quantity = quantity + %s, damaged = damaged + %s WHERE id=%s",
            (returned_quantity, damaged_quantity, b['consumable_id'])
        )
        connection.commit()
        flash('Return recorded and stock updated', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Return error: {err}")
        flash('Error recording return', 'error')
    finally:
        cursor.close()
        connection.close()
    return redirect(url_for('admin_consumables'))


@app.route('/admin/consumables/borrower_suggest')
@admin_required
def admin_borrower_suggest():
    q = (request.args.get('q') or '').strip()
    connection = get_db_connection()
    if not connection:
        return jsonify([])
    cursor = connection.cursor()
    try:
        if q:
            cursor.execute("SELECT DISTINCT borrower_name FROM consumable_borrows WHERE borrower_name LIKE %s ORDER BY borrower_name LIMIT 10", (f'%{q}%',))
        else:
            cursor.execute("SELECT DISTINCT borrower_name FROM consumable_borrows ORDER BY borrower_name LIMIT 10")
        names = [row[0] for row in cursor.fetchall()]
    finally:
        cursor.close()
        connection.close()
    return jsonify(names)


@app.route('/admin/labs/add', methods=['POST'])
@admin_required
def admin_add_lab():
    """Add a new laboratory (name, status)"""
    name = (request.form.get('name') or '').strip()
    status = (request.form.get('status') or '').strip()

    if not name or status not in ('Active', 'Inactive', 'Maintenance'):
        flash('Please provide a valid Lab Name and Status', 'error')
        return redirect(url_for('admin_inventory'))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_inventory'))

    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO laboratory (name, status)
            VALUES (%s, %s)
            """,
            (name, status),
        )
        connection.commit()
        log_admin_action('Add Lab', f'Added lab: {name} ({status})')
        flash('Laboratory added successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Add laboratory error: {err}")
        flash('Error adding laboratory', 'error')
    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('admin_inventory'))

@app.route('/admin/inventory/add', methods=['GET', 'POST'])
@admin_required
def admin_add_consumable():
    """Add new consumable"""
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        category = request.form.get('category')
        quantity = int(request.form.get('quantity', 0))
        image_url = request.form.get('image_url', '')
        
        if not all([name, description, category]):
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('admin_inventory'))
        
        connection = get_db_connection()
        if not connection:
            flash('Database connection error', 'error')
            return redirect(url_for('admin_inventory'))
        
        cursor = connection.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO consumables (name, description, category, quantity, image_url) 
                VALUES (%s, %s, %s, %s, %s)
            """, (name, description, category, quantity, image_url))
            
            connection.commit()
            log_admin_action('Add Consumable', f'Added: {name}')
            flash('Consumable added successfully!', 'success')
            return redirect(url_for('admin_inventory'))
            
        except mysql.connector.Error as err:
            connection.rollback()
            flash('Error adding consumable', 'error')
            logger.error(f"Add consumable error: {err}")
        finally:
            cursor.close()
            connection.close()
    
    # For GET requests, redirect to inventory (no standalone page)
    return redirect(url_for('admin_inventory'))

@app.route('/admin/inventory/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_consumable(id):
    """Edit existing consumable"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_inventory'))
    
    cursor = connection.cursor(dictionary=True)
    
    if request.method == 'GET':
        cursor.execute("SELECT * FROM consumables WHERE id = %s", (id,))
        consumable = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if not consumable:
            flash('Consumable not found', 'error')
            return redirect(url_for('admin_inventory'))
        
        return render_template('admin/edit_consumable.html', consumable=consumable)
    
    # Process edit
    name = request.form.get('name')
    description = request.form.get('description')
    category = request.form.get('category')
    quantity = int(request.form.get('quantity', 0))
    image_url = request.form.get('image_url', '')
    
    if not all([name, description, category]):
        flash('Please fill in all required fields', 'error')
        return redirect(url_for('admin_edit_consumable', id=id))
    
    cursor = connection.cursor()
    
    try:
        cursor.execute("""
            UPDATE consumables 
            SET name = %s, description = %s, category = %s, quantity = %s, image_url = %s 
            WHERE id = %s
        """, (name, description, category, quantity, image_url, id))
        
        connection.commit()
        log_admin_action('Edit Consumable', f'Edited: {name}')
        flash('Consumable updated successfully!', 'success')
        return redirect(url_for('admin_inventory'))
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash('Error updating consumable', 'error')
        logger.error(f"Edit consumable error: {err}")
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_edit_consumable', id=id))

@app.route('/admin/inventory/delete/<int:id>', methods=['POST'])
@admin_required
def admin_delete_consumable(id):
    """Delete consumable"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_inventory'))
    
    cursor = connection.cursor(dictionary=True)
    
    # Get consumable name for audit log
    cursor.execute("SELECT name FROM consumables WHERE id = %s", (id,))
    consumable = cursor.fetchone()
    
    if not consumable:
        flash('Consumable not found', 'error')
        return redirect(url_for('admin_inventory'))
    
    cursor = connection.cursor()
    
    try:
        cursor.execute("DELETE FROM consumables WHERE id = %s", (id,))
        connection.commit()
        log_admin_action('Delete Consumable', f'Deleted: {consumable["name"]}')
        flash('Consumable deleted successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        flash('Error deleting consumable', 'error')
        logger.error(f"Delete consumable error: {err}")
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_inventory'))

@app.route('/admin/orders')
@admin_required
def admin_orders():
    """View and manage orders"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('admin/orders.html', orders=[])
    
    cursor = connection.cursor(dictionary=True)
    
    # Get filter parameters
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    # Build query
    query = """
        SELECT o.*, COUNT(oi.id) as item_count, 
               SUM(oi.quantity) as total_quantity
        FROM orders o 
        LEFT JOIN order_items oi ON o.id = oi.order_id 
        WHERE 1=1
    """
    params = []
    
    if status:
        query += " AND o.status = %s"
        params.append(status)
    
    if search:
        query += " AND (o.user_name LIKE %s OR o.department LIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    query += " GROUP BY o.id ORDER BY o.created_at DESC"
    
    cursor.execute(query, params)
    orders = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return render_template('admin/orders.html',
                         orders=orders,
                         status=status,
                         search=search)

@app.route('/admin/orders/<int:id>')
@admin_required
def admin_order_detail(id):
    """View order details"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_orders'))
    
    cursor = connection.cursor(dictionary=True)
    
    # Get order details
    cursor.execute("SELECT * FROM orders WHERE id = %s", (id,))
    order = cursor.fetchone()
    
    if not order:
        flash('Order not found', 'error')
        return redirect(url_for('admin_orders'))
    
    # Get order items
    cursor.execute("""
        SELECT oi.*, c.name, c.description, c.category, c.quantity as stock_quantity
        FROM order_items oi
        JOIN consumables c ON oi.consumable_id = c.id
        WHERE oi.order_id = %s
    """, (id,))
    order_items = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return render_template('admin/order_detail.html', order=order, order_items=order_items)

@app.route('/admin/orders/<int:id>/approve', methods=['POST'])
@admin_required
def admin_approve_order(id):
    """Approve order and reduce stock"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_order_detail', id=id))
    
    cursor = connection.cursor()
    
    try:
        # Update order status
        cursor.execute("UPDATE orders SET status = 'Approved' WHERE id = %s", (id,))
        
        # Get order items and reduce stock
        cursor.execute("""
            SELECT oi.consumable_id, oi.quantity, c.name, c.quantity as current_stock
            FROM order_items oi
            JOIN consumables c ON oi.consumable_id = c.id
            WHERE oi.order_id = %s
        """, (id,))
        
        order_items = cursor.fetchall()
        
        for item in order_items:
            consumable_id, quantity, name, current_stock = item
            new_stock = current_stock - quantity
            
            if new_stock < 0:
                connection.rollback()
                flash(f'Insufficient stock for {name}. Available: {current_stock}, Requested: {quantity}', 'error')
                return redirect(url_for('admin_order_detail', id=id))
            
            cursor.execute("""
                UPDATE consumables SET quantity = %s WHERE id = %s
            """, (new_stock, consumable_id))
        
        connection.commit()
        log_admin_action('Approve Order', f'Approved order #{id}')
        flash('Order approved successfully! Stock quantities updated.', 'success')
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash('Error approving order', 'error')
        logger.error(f"Approve order error: {err}")
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_order_detail', id=id))

@app.route('/admin/orders/<int:id>/reject', methods=['POST'])
@admin_required
def admin_reject_order(id):
    """Reject order"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_order_detail', id=id))
    
    cursor = connection.cursor()
    
    try:
        cursor.execute("UPDATE orders SET status = 'Rejected' WHERE id = %s", (id,))
        connection.commit()
        log_admin_action('Reject Order', f'Rejected order #{id}')
        flash('Order rejected successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        flash('Error rejecting order', 'error')
        logger.error(f"Reject order error: {err}")
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_order_detail', id=id))

@app.route('/admin/export/orders')
@admin_required
def admin_export_orders():
    """Export orders to CSV"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_orders'))
    
    cursor = connection.cursor(dictionary=True)
    
    # Get all orders with items
    cursor.execute("""
        SELECT o.*, GROUP_CONCAT(CONCAT(c.name, ' (', oi.quantity, ')') SEPARATOR '; ') as items
        FROM orders o 
        LEFT JOIN order_items oi ON o.id = oi.order_id 
        LEFT JOIN consumables c ON oi.consumable_id = c.id
        GROUP BY o.id 
        ORDER BY o.created_at DESC
    """)
    
    orders = cursor.fetchall()
    cursor.close()
    connection.close()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Order ID', 'User Name', 'Department', 'Purpose', 'Date Needed', 'Status', 'Items', 'Created At'])
    
    # Write data
    for order in orders:
        writer.writerow([
            order['id'],
            order['user_name'],
            order['department'],
            order['purpose'],
            order['date_needed'],
            order['status'],
            order['items'] or 'No items',
            order['created_at']
        ])
    
    output.seek(0)
    
    log_admin_action('Export Orders', f'Exported {len(orders)} orders to CSV')
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'orders_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@app.route('/admin/export/inventory')
@admin_required
def admin_export_inventory():
    """Export inventory to Excel"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_inventory'))
    
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT * FROM consumables ORDER BY category, name")
    consumables = cursor.fetchall()
    cursor.close()
    connection.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write headers
    headers = ['ID', 'Name', 'Description', 'Category', 'Quantity', 'Image URL', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row, item in enumerate(consumables, 2):
        ws.cell(row=row, column=1, value=item['id'])
        ws.cell(row=row, column=2, value=item['name'])
        ws.cell(row=row, column=3, value=item['description'])
        ws.cell(row=row, column=4, value=item['category'])
        ws.cell(row=row, column=5, value=item['quantity'])
        ws.cell(row=row, column=6, value=item['image_url'])
        ws.cell(row=row, column=7, value=item['created_at'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_admin_action('Export Inventory', f'Exported {len(consumables)} items to Excel')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventory_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/audit-logs')
@admin_required
def admin_audit_logs():
    """View audit logs"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return render_template('admin/audit_logs.html', logs=[])
    
    cursor = connection.cursor(dictionary=True)
    
    # Get logs with pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    cursor.execute("SELECT COUNT(*) as total FROM audit_logs")
    total_logs = cursor.fetchone()['total']
    
    cursor.execute("""
        SELECT * FROM audit_logs 
        ORDER BY timestamp DESC 
        LIMIT %s OFFSET %s
    """, (per_page, (page - 1) * per_page))
    
    logs = cursor.fetchall()
    cursor.close()
    connection.close()
    
    total_pages = (total_logs + per_page - 1) // per_page
    
    return render_template('admin/audit_logs.html',
                         logs=logs,
                         page=page,
                         total_pages=total_pages)


@app.route('/admin/categories/add', methods=['POST'])
@admin_required
def admin_add_category():
    """Add a new asset category via prompt"""
    name = (request.form.get('name') or '').strip()

    if not name:
        flash('Please provide a category name', 'error')
        return redirect(request.referrer or url_for('admin_inventory'))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(request.referrer or url_for('admin_inventory'))

    cursor = connection.cursor()
    try:
        # Insert category if not exists
        cursor.execute("SELECT id FROM asset_categories WHERE name = %s", (name,))
        row = cursor.fetchone()
        if not row:
            cursor.execute("INSERT INTO asset_categories (name) VALUES (%s)", (name,))
            connection.commit()
            log_admin_action('Add Category', f'Added category: {name}')
            flash('Category added successfully!', 'success')
        else:
            flash('Category already exists', 'info')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Add category error: {err}")
        flash('Error adding category', 'error')
    finally:
        cursor.close()
        connection.close()

    return redirect(request.referrer or url_for('admin_inventory'))

@app.route('/admin/labs/<int:lab_id>/assets')
@admin_required
def admin_lab_assets(lab_id):
    """View all assets in a specific laboratory"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_inventory'))
    
    cursor = connection.cursor(dictionary=True)
    
    try:
        # Get lab information
        cursor.execute("SELECT * FROM laboratory WHERE id = %s", (lab_id,))
        lab = cursor.fetchone()
        
        if not lab:
            flash('Laboratory not found', 'error')
            return redirect(url_for('admin_inventory'))
        
        # Get search, filter, sort, and pagination parameters
        search = request.args.get('search', '')
        category_filter = request.args.get('category_filter', '')
        status_filter = request.args.get('status_filter', '')
        sort_by = request.args.get('sort_by', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 15, type=int)
        
        # Build query for assets
        query = "SELECT * FROM lab_assets WHERE lab_id = %s"
        params = [lab_id]
        
        if search:
            query += " AND (name LIKE %s OR asset_code LIKE %s OR description LIKE %s)"
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
        
        if category_filter:
            query += " AND category = %s"
            params.append(category_filter)
        
        if status_filter:
            query += " AND status = %s"
            params.append(status_filter)
        
        # Sorting (allowlist to avoid SQL injection)
        allowed_sort_columns = {
            'name': 'name',
            'asset_code': 'asset_code',
            'category': 'category',
            'status': 'status',
            'stock_date': 'stock_date',
            'updated_at': 'updated_at',
            'created_at': 'created_at',
        }
        sort_col = allowed_sort_columns.get(sort_by, 'created_at')
        sort_dir = 'ASC' if str(sort_order).lower() == 'asc' else 'DESC'

        # Count for pagination
        count_query = query.replace('SELECT *', 'SELECT COUNT(*) AS cnt')
        cursor.execute(count_query, params)
        total_assets = cursor.fetchone()['cnt']

        # Add order and pagination
        query += f" ORDER BY {sort_col} {sort_dir} LIMIT %s OFFSET %s"
        params.extend([per_page, (page - 1) * per_page])

        cursor.execute(query, params)
        assets = cursor.fetchall()
        
        # Load categories for filters and add modal
        cursor.execute("SELECT name FROM asset_categories ORDER BY name ASC")
        categories = [row['name'] for row in cursor.fetchall()]
        
    finally:
        cursor.close()
        connection.close()
    
    total_pages = (total_assets + per_page - 1) // per_page

    return render_template('admin/lab_assets.html', 
                         lab=lab, 
                         assets=assets,
                         search=search,
                         category_filter=category_filter,
                         status_filter=status_filter,
                         categories=categories,
                         sort_by=sort_by,
                         sort_order=sort_order,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages)


@app.route('/admin/labs/<int:lab_id>/assets/add', methods=['POST'])
@admin_required
def admin_add_asset(lab_id):
    """Add a new asset to a laboratory"""
    name = (request.form.get('name') or '').strip()
    asset_code = (request.form.get('asset_code') or '').strip()
    category = (request.form.get('category') or '').strip()
    status = (request.form.get('status') or '').strip()
    purchase_date = request.form.get('purchase_date') or None
    description = (request.form.get('description') or '').strip()
    
    if not all([name, category, status]):
        flash('Please provide Asset Name, Category, and Status', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    
    if status not in ('Available', 'In Use', 'Maintenance', 'Retired', 'Damaged'):
        flash('Invalid status selected', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    
    cursor = connection.cursor()
    try:
        # Ensure category exists in asset_categories
        cursor.execute("SELECT id FROM asset_categories WHERE name = %s", (category,))
        category_row = cursor.fetchone()
        if not category_row:
            cursor.execute("INSERT INTO asset_categories (name) VALUES (%s)", (category,))

        cursor.execute(
            """
            INSERT INTO lab_assets (lab_id, name, asset_code, category, status, stock_date, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (lab_id, name, asset_code or None, category, status, purchase_date, description),
        )
        connection.commit()
        log_admin_action('Add Asset', f'Added asset: {name} to lab #{lab_id}')
        flash('Asset added successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Add asset error: {err}")
        flash('Error adding asset', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_lab_assets', lab_id=lab_id))


@app.route('/admin/labs/<int:lab_id>/assets/import', methods=['POST'])
@admin_required
def admin_import_assets(lab_id):
    """Import assets from an Excel file and skip duplicates by asset_code"""
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Please choose an Excel (.xlsx) file to upload', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Invalid file type. Please upload a .xlsx Excel file', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Excel load error: {e}")
        flash('Unable to read the Excel file. Please check the format.', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    # Expected headers (case-insensitive match)
    expected_headers = {
        'asset name': 'name',
        'asset code': 'asset_code',
        'category': 'category',
        'status': 'status',
        'stock date': 'stock_date',
        'description': 'description',
    }

    # Read header row and map columns
    header_cells = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    header_map = {}
    for idx, header in enumerate(header_cells):
        key = header.lower()
        if key in expected_headers:
            header_map[expected_headers[key]] = idx

    # Validate required minimal headers
    for req in ('name', 'category', 'status'):
        if req not in header_map:
            flash('Missing column in Excel: ' + req.replace('_', ' ').title(), 'error')
            return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    # Read rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(col):
            if col not in header_map:
                return ''
            val = row[header_map[col]]
            return '' if val is None else str(val).strip()

        name = get('name')
        asset_code = get('asset_code')
        category = get('category')
        status = get('status')
        stock_date_raw = row[header_map['stock_date']] if 'stock_date' in header_map else None
        description = get('description')

        if not name or not category or not status:
            continue

        # Normalize status
        valid_status = {'Available', 'In Use', 'Maintenance', 'Retired', 'Damaged'}
        if status not in valid_status:
            # Attempt to normalize common variants
            s = status.strip().lower()
            if s in ('available',):
                status = 'Available'
            elif s in ('in use', 'in-use', 'in_use'):
                status = 'In Use'
            elif s in ('maintenance', 'maint'):
                status = 'Maintenance'
            elif s in ('retired', 'inactive'):
                status = 'Retired'
            elif s in ('damaged', 'broken'):
                status = 'Damaged'
            else:
                # default to Available if unrecognized
                status = 'Available'

        # Parse date to YYYY-MM-DD if possible
        stock_date = None
        if stock_date_raw:
            try:
                if hasattr(stock_date_raw, 'strftime'):
                    stock_date = stock_date_raw.strftime('%Y-%m-%d')
                else:
                    # try parse as string YYYY-MM-DD or DD/MM/YYYY
                    txt = str(stock_date_raw).strip().replace('/', '-')
                    parts = txt.split('-')
                    if len(parts) == 3:
                        if len(parts[0]) == 4:
                            stock_date = f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
                        else:
                            stock_date = f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            except Exception:
                stock_date = None

        rows.append({
            'name': name,
            'asset_code': asset_code or None,
            'category': category,
            'status': status,
            'stock_date': stock_date,
            'description': description,
        })

    if not rows:
        flash('No valid rows found in the Excel file', 'info')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor(dictionary=True)
    inserted = 0
    skipped_codes = []
    try:
        # Ensure all categories exist
        categories_needed = sorted({r['category'] for r in rows})
        for cat in categories_needed:
            cursor.execute("SELECT id FROM asset_categories WHERE name = %s", (cat,))
            if not cursor.fetchone():
                cursor.execute("INSERT INTO asset_categories (name) VALUES (%s)", (cat,))

        # Preload existing asset codes for this lab
        codes = [r['asset_code'] for r in rows if r['asset_code']]
        existing_codes = set()
        if codes:
            format_strings = ','.join(['%s'] * len(codes))
            cursor.execute(f"SELECT asset_code FROM lab_assets WHERE lab_id = %s AND asset_code IN ({format_strings})", [lab_id, *codes])
            existing_codes = {row['asset_code'] for row in cursor.fetchall() if row['asset_code']}

        # Track duplicates in file itself
        seen_codes = set()

        for r in rows:
            code = r['asset_code']
            if code:
                if code in existing_codes or code in seen_codes:
                    skipped_codes.append(code)
                    continue
                seen_codes.add(code)

            cursor.execute(
                """
                INSERT INTO lab_assets (lab_id, name, asset_code, category, status, stock_date, description)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (lab_id, r['name'], r['asset_code'], r['category'], r['status'], r['stock_date'], r['description'])
            )
            inserted += 1

        connection.commit()
        msg = f"Imported {inserted} asset(s)."
        if skipped_codes:
            unique_skipped = sorted(set(skipped_codes))
            msg += f" Skipped {len(unique_skipped)} duplicate Asset Code(s): {', '.join(unique_skipped[:10])}"
            if len(unique_skipped) > 10:
                msg += " ..."
        flash(msg, 'success')
        log_admin_action('Import Assets', f'Imported {inserted} assets to lab #{lab_id}; skipped {len(set(skipped_codes))} duplicates')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Import assets DB error: {err}")
        flash('Error importing assets. Please check the file and try again.', 'error')
    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('admin_lab_assets', lab_id=lab_id))


@app.route('/admin/labs/<int:lab_id>/assets/template')
@admin_required
def admin_assets_template(lab_id):
    """Download a simple Excel template for importing assets"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = [
        'Asset Name',
        'Asset Code',
        'Category',
        'Status',
        'Stock Date',
        'Description',
    ]

    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Provide a sample row (optional)
    ws.cell(row=2, column=1, value='Laptop Lenovo T14')
    ws.cell(row=2, column=2, value='IPRC-T/LIB/CU373')
    ws.cell(row=2, column=3, value='Computer Equipment')
    ws.cell(row=2, column=4, value='Available')
    ws.cell(row=2, column=5, value='2025-01-01')
    ws.cell(row=2, column=6, value='14 inch laptop for lab use')

    # Auto width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_admin_action('Download Assets Template', f'Lab #{lab_id} template downloaded')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='assets_import_template.xlsx'
    )


def _fetch_assets_for_export(cursor, lab_id, request_args, selected_ids=None):
    """Helper to fetch assets by selected IDs or current filters."""
    params = [lab_id]
    query = "SELECT * FROM lab_assets WHERE lab_id = %s"

    # If explicit IDs provided, take precedence
    if selected_ids:
        placeholders = ','.join(['%s'] * len(selected_ids))
        query += f" AND id IN ({placeholders})"
        params.extend(selected_ids)
    else:
        search = request_args.get('search', '')
        category_filter = request_args.get('category_filter', '')
        status_filter = request_args.get('status_filter', '')
        if search:
            query += " AND (name LIKE %s OR asset_code LIKE %s OR description LIKE %s)"
            like = f"%{search}%"
            params.extend([like, like, like])
        if category_filter:
            query += " AND category = %s"
            params.append(category_filter)
        if status_filter:
            query += " AND status = %s"
            params.append(status_filter)

    query += " ORDER BY created_at DESC"
    cursor.execute(query, params)
    return cursor.fetchall()


@app.route('/admin/labs/<int:lab_id>/assets/export/excel')
@admin_required
def admin_export_assets_excel(lab_id):
    """Export selected or filtered assets to Excel"""
    # Parse selected IDs from query
    asset_ids = request.args.getlist('asset_ids', type=int)

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor(dictionary=True)
    try:
        assets = _fetch_assets_for_export(cursor, lab_id, request.args, asset_ids or None)
    finally:
        cursor.close()
        connection.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = ['Asset ID', 'Asset Name', 'Asset Code', 'Category', 'Status', 'Stock Date', 'Description', 'Created At', 'Updated At']
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, a in enumerate(assets, start=2):
        ws.cell(row=row_idx, column=1, value=a['id'])
        ws.cell(row=row_idx, column=2, value=a['name'])
        ws.cell(row=row_idx, column=3, value=a.get('asset_code'))
        ws.cell(row=row_idx, column=4, value=a['category'])
        ws.cell(row=row_idx, column=5, value=a['status'])
        ws.cell(row=row_idx, column=6, value=a.get('stock_date'))
        ws.cell(row=row_idx, column=7, value=a.get('description'))
        ws.cell(row=row_idx, column=8, value=a.get('created_at'))
        ws.cell(row=row_idx, column=9, value=a.get('updated_at'))

    # Auto size
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_admin_action('Export Assets Excel', f'Lab #{lab_id} exported {len(assets)} assets to Excel')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'lab_{lab_id}_assets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/admin/labs/<int:lab_id>/assets/export/pdf')
@admin_required
def admin_export_assets_pdf(lab_id):
    """Export selected or filtered assets to a simple PDF table"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        flash('PDF export requires reportlab. Please install it or use Excel export.', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    asset_ids = request.args.getlist('asset_ids', type=int)

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor(dictionary=True)
    try:
        assets = _fetch_assets_for_export(cursor, lab_id, request.args, asset_ids or None)
    finally:
        cursor.close()
        connection.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    data = [[
        'Asset ID', 'Asset Name', 'Asset Code', 'Category', 'Status', 'Stock Date', 'Description'
    ]]
    for a in assets:
        data.append([
            str(a['id']), a['name'], a.get('asset_code') or '', a['category'], a['status'],
            str(a.get('stock_date') or ''), a.get('description') or ''
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))

    doc.build([Paragraph(f"Lab {lab_id} Assets Export ({len(assets)} items)", styles['Heading2']), table])
    buffer.seek(0)

    log_admin_action('Export Assets PDF', f'Lab #{lab_id} exported {len(assets)} assets to PDF')
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'lab_{lab_id}_assets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@app.route('/admin/labs/<int:lab_id>/assets/<int:asset_id>/edit', methods=['POST'])
@admin_required
def admin_edit_asset(lab_id, asset_id):
    """Edit an existing asset"""
    name = (request.form.get('name') or '').strip()
    asset_code = (request.form.get('asset_code') or '').strip()
    category = (request.form.get('category') or '').strip()
    status = (request.form.get('status') or '').strip()
    stock_date = request.form.get('stock_date') or None
    description = (request.form.get('description') or '').strip()

    if not all([name, category, status]):
        flash('Please provide Asset Name, Category, and Status', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    if status not in ('Available', 'In Use', 'Maintenance', 'Retired', 'Damaged'):
        flash('Invalid status selected', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor()
    try:
        # Ensure category exists
        cursor.execute("SELECT id FROM asset_categories WHERE name = %s", (category,))
        category_row = cursor.fetchone()
        if not category_row:
            cursor.execute("INSERT INTO asset_categories (name) VALUES (%s)", (category,))

        cursor.execute(
            """
            UPDATE lab_assets
            SET name=%s, asset_code=%s, category=%s, status=%s, stock_date=%s, description=%s
            WHERE id=%s AND lab_id=%s
            """,
            (name, asset_code or None, category, status, stock_date, description, asset_id, lab_id),
        )
        connection.commit()
        log_admin_action('Edit Asset', f'Edited asset #{asset_id} in lab #{lab_id}')
        flash('Asset updated successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Edit asset error: {err}")
        flash('Error updating asset', 'error')
    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('admin_lab_assets', lab_id=lab_id))


@app.route('/admin/labs/<int:lab_id>/assets/<int:asset_id>/delete', methods=['POST'])
@admin_required
def admin_delete_asset(lab_id, asset_id):
    """Delete an asset after confirmation"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM lab_assets WHERE id = %s AND lab_id = %s", (asset_id, lab_id))
        connection.commit()
        log_admin_action('Delete Asset', f'Deleted asset #{asset_id} from lab #{lab_id}')
        flash('Asset deleted successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Delete asset error: {err}")
        flash('Error deleting asset', 'error')
    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('admin_lab_assets', lab_id=lab_id))

@app.route('/admin/labs/<int:lab_id>/edit', methods=['POST'])
@admin_required
def admin_edit_lab(lab_id):
    """Edit laboratory information"""
    name = (request.form.get('name') or '').strip()
    status = (request.form.get('status') or '').strip()
    
    if not name or status not in ('Active', 'Inactive', 'Maintenance'):
        flash('Please provide a valid Lab Name and Status', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            UPDATE laboratory SET name = %s, status = %s WHERE id = %s
            """,
            (name, status, lab_id),
        )
        connection.commit()
        log_admin_action('Edit Lab', f'Updated lab: {name} (ID: {lab_id})')
        flash('Laboratory updated successfully!', 'success')
    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Edit laboratory error: {err}")
        flash('Error updating laboratory', 'error')
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('admin_lab_assets', lab_id=lab_id))


@app.route('/admin/labs/<int:lab_id>/delete', methods=['POST'])
@admin_required
def admin_delete_lab(lab_id):
    """Delete a laboratory only if it has no assets"""
    connection = get_db_connection()
    if not connection:
        flash('Database connection error', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))

    cursor = connection.cursor(dictionary=True)
    try:
        # Verify lab exists and get its name
        cursor.execute("SELECT name FROM laboratory WHERE id = %s", (lab_id,))
        lab = cursor.fetchone()
        if not lab:
            flash('Laboratory not found', 'error')
            return redirect(url_for('admin_inventory'))

        # Check if lab has any assets
        cursor.execute("SELECT COUNT(*) AS cnt FROM lab_assets WHERE lab_id = %s", (lab_id,))
        asset_count = cursor.fetchone()['cnt']

        if asset_count and int(asset_count) > 0:
            flash('Cannot delete this laboratory because it has assets. Remove all assets first.', 'error')
            return redirect(url_for('admin_lab_assets', lab_id=lab_id))

        # Safe to delete
        cursor.execute("DELETE FROM laboratory WHERE id = %s", (lab_id,))
        connection.commit()
        log_admin_action('Delete Lab', f"Deleted lab: {lab['name']} (ID: {lab_id})")
        flash('Laboratory deleted successfully!', 'success')
        return redirect(url_for('admin_inventory'))

    except mysql.connector.Error as err:
        connection.rollback()
        logger.error(f"Delete laboratory error: {err}")
        flash('Error deleting laboratory', 'error')
        return redirect(url_for('admin_lab_assets', lab_id=lab_id))
    finally:
        cursor.close()
        connection.close()

if __name__ == '__main__':
    # Initialize database
    if init_database():
        print("Database initialized successfully!")
    else:
        print("Database initialization failed!")
    
    app.run(debug=True, host='0.0.0.0', port=5000) 